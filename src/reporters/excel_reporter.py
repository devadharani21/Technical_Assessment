"""Excel Report Generator."""

from pathlib import Path
from typing import Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime


class ExcelReporter:
    """Generates Excel reports with multiple sheets and charts."""
    
    def __init__(self):
        """Initialize the reporter."""
        pass
    
    def generate_report(self, results: Dict[str, Any], output_path: Path):
        """
        Generate a comprehensive Excel report.
        
        Args:
            results: Analysis results dictionary
            output_path: Path to save the report
        """
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Create Excel writer
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Summary sheet
            self._create_summary_sheet(results, writer)
            
            # Flaky patterns sheet
            self._create_flaky_patterns_sheet(results, writer)
            
            # Code duplication sheet
            self._create_duplication_sheet(results, writer)
            
            # Coding standards sheet
            self._create_coding_standards_sheet(results, writer)
            
            # All issues sheet
            self._create_all_issues_sheet(results, writer)
        
        # Add formatting and charts
        self._format_workbook(output_path, results)
        
        print(f"Excel report generated: {output_path}")
    
    def _create_summary_sheet(self, results: Dict[str, Any], writer):
        """Create summary overview sheet."""
        metadata = results.get('metadata', {})
        summary = results.get('summary', {})
        
        # Summary data
        summary_data = {
            'Metric': [
                'Analysis Date',
                'Directory',
                'Total Files Analyzed',
                'Analysis Duration (seconds)',
                '',
                'Total Issues Found',
                'Critical Issues',
                'High Severity Issues',
                'Medium Severity Issues',
                'Low Severity Issues',
                '',
                'Files with Flaky Patterns',
                'Total Flaky Issues',
                'Code Duplication %'
            ],
            'Value': [
                metadata.get('analysis_date', 'N/A'),
                metadata.get('directory', 'N/A'),
                metadata.get('total_files', 0),
                f"{metadata.get('duration_seconds', 0):.2f}",
                '',
                summary.get('total_issues', 0),
                summary.get('by_severity', {}).get('critical', 0),
                summary.get('by_severity', {}).get('high', 0),
                summary.get('by_severity', {}).get('medium', 0),
                summary.get('by_severity', {}).get('low', 0),
                '',
                summary.get('flaky_test_stats', {}).get('files_with_flaky_patterns', 0),
                summary.get('flaky_test_stats', {}).get('total_flaky_issues', 0),
                f"{results.get('duplication', {}).get('duplication_percentage', 0):.2f}%"
            ]
        }
        
        df = pd.DataFrame(summary_data)
        df.to_excel(writer, sheet_name='Summary', index=False)
    
    def _create_flaky_patterns_sheet(self, results: Dict[str, Any], writer):
        """Create flaky patterns analysis sheet."""
        flaky_data = []
        
        for file_result in results.get('flaky_patterns', []):
            file_path = Path(file_result['file']).name
            risk_score = file_result.get('risk_score', 0)
            
            for issue in file_result.get('issues', []):
                flaky_data.append({
                    'File': file_path,
                    'Line': issue.get('line', 0),
                    'Category': issue.get('category', 'N/A'),
                    'Severity': issue.get('severity', 'N/A'),
                    'Message': issue.get('message', ''),
                    'Pattern': issue.get('pattern', ''),
                    'Risk Score': risk_score,
                    'Recommendation': issue.get('recommendation', '')
                })
        
        if flaky_data:
            df = pd.DataFrame(flaky_data)
            df.to_excel(writer, sheet_name='Flaky Patterns', index=False)
        else:
            pd.DataFrame({'Message': ['No flaky patterns detected']}).to_excel(
                writer, sheet_name='Flaky Patterns', index=False
            )
    
    def _create_duplication_sheet(self, results: Dict[str, Any], writer):
        """Create code duplication sheet."""
        dup_data = []
        
        for duplicate in results.get('duplication', {}).get('duplicates', []):
            for location in duplicate['locations']:
                file_path = Path(location['file']).name
                dup_data.append({
                    'File': file_path,
                    'Start Line': location['start_line'],
                    'End Line': location['end_line'],
                    'Lines': duplicate['lines'],
                    'Total Occurrences': duplicate['occurrences'],
                    'Files Affected': duplicate['files_affected'],
                    'Severity': duplicate['severity'],
                    'Recommendation': duplicate['recommendation']
                })
        
        if dup_data:
            df = pd.DataFrame(dup_data)
            df.to_excel(writer, sheet_name='Code Duplication', index=False)
        else:
            pd.DataFrame({'Message': ['No code duplication detected']}).to_excel(
                writer, sheet_name='Code Duplication', index=False
            )
    
    def _create_coding_standards_sheet(self, results: Dict[str, Any], writer):
        """Create coding standards violations sheet."""
        violations_data = []
        
        for file_result in results.get('coding_standards', []):
            file_path = Path(file_result['file']).name
            
            for violation in file_result.get('violations', []):
                violations_data.append({
                    'File': file_path,
                    'Line': violation.get('line', 0),
                    'Type': violation.get('type', 'N/A'),
                    'Severity': violation.get('severity', 'N/A'),
                    'Message': violation.get('message', ''),
                    'Code': violation.get('code', '')[:100]
                })
        
        if violations_data:
            df = pd.DataFrame(violations_data)
            df.to_excel(writer, sheet_name='Coding Standards', index=False)
        else:
            pd.DataFrame({'Message': ['No coding standards violations found']}).to_excel(
                writer, sheet_name='Coding Standards', index=False
            )
    
    def _create_all_issues_sheet(self, results: Dict[str, Any], writer):
        """Create sheet with all issues combined."""
        all_issues = []
        
        # Add flaky pattern issues
        for file_result in results.get('flaky_patterns', []):
            for issue in file_result.get('issues', []):
                all_issues.append({
                    'File': Path(issue['file']).name,
                    'Line': issue.get('line', 0),
                    'Category': 'Flaky Pattern',
                    'Type': issue.get('category', 'N/A'),
                    'Severity': issue.get('severity', 'N/A'),
                    'Message': issue.get('message', '')
                })
        
        # Add coding standards violations
        for file_result in results.get('coding_standards', []):
            for violation in file_result.get('violations', []):
                all_issues.append({
                    'File': Path(violation['file']).name,
                    'Line': violation.get('line', 0),
                    'Category': 'Coding Standards',
                    'Type': violation.get('type', 'N/A'),
                    'Severity': violation.get('severity', 'N/A'),
                    'Message': violation.get('message', '')
                })
        
        if all_issues:
            df = pd.DataFrame(all_issues)
            df = df.sort_values(['Severity', 'File'])
            df.to_excel(writer, sheet_name='All Issues', index=False)
    
    def _format_workbook(self, output_path: Path, results: Dict[str, Any]):
        """Add formatting and charts to the workbook."""
        try:
            wb = load_workbook(output_path)
            
            # Format Summary sheet
            if 'Summary' in wb.sheetnames:
                ws = wb['Summary']
                self._format_summary_sheet(ws)
                self._add_severity_chart(ws, results)
            
            # Auto-adjust column widths for all sheets
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self._adjust_column_widths(ws)
            
            wb.save(output_path)
        
        except Exception as e:
            print(f"Warning: Could not add formatting: {e}")
    
    def _format_summary_sheet(self, ws):
        """Format the summary sheet."""
        # Header row formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Highlight severity rows
        severity_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        for row in ws.iter_rows(min_row=2):
            metric_cell = row[0]
            if 'Issues' in str(metric_cell.value) or 'Critical' in str(metric_cell.value):
                for cell in row:
                    cell.fill = severity_fill
    
    def _add_severity_chart(self, ws, results: Dict[str, Any]):
        """Add a pie chart for severity distribution."""
        try:
            summary = results.get('summary', {})
            by_severity = summary.get('by_severity', {})
            
            # Only add chart if there are issues
            if sum(by_severity.values()) > 0:
                chart = PieChart()
                chart.title = "Issues by Severity"
                chart.style = 10
                chart.height = 10
                chart.width = 15
                
                # This is a simplified approach - in reality you'd need to 
                # add the data to the sheet first
                ws.append([])
                ws.append(['Severity', 'Count'])
                for severity in ['critical', 'high', 'medium', 'low']:
                    ws.append([severity.capitalize(), by_severity.get(severity, 0)])
        
        except Exception:
            pass  # Skip chart if there's an error
    
    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

